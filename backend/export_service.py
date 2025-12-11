# Excel Export Service

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
import config

class ExportService:
    """Export exam results to Excel"""
    
    def __init__(self, storage):
        self.storage = storage
    
    def export_to_excel(self, exam_id: str) -> Path:
        """Export exam results to Excel file"""
        exam = self.storage.load_exam(exam_id)
        if not exam:
            raise Exception("Exam not found")
        
        results = self.storage.list_results_by_exam(exam_id)
        stats = self.storage.get_exam_statistics(exam_id)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb, exam, stats)
        self._create_scores_sheet(wb, exam, results)
        self._create_details_sheet(wb, exam, results)
        
        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_{exam_id}_{timestamp}.xlsx"
        file_path = config.EXPORTS_DIR / filename
        
        try:
            # Ensure directory exists
            config.EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
            
            # Save with explicit engine
            wb.save(str(file_path))
            print(f"✓ Excel exported to: {file_path}")
            
            # Verify file exists and has size
            if file_path.exists() and file_path.stat().st_size > 0:
                return file_path
            else:
                raise Exception("Excel file was not created properly")
                
        except Exception as e:
            print(f"Error saving Excel: {e}")
            raise Exception(f"Failed to save Excel file: {str(e)}")
    
    def _create_summary_sheet(self, wb, exam, stats):
        """Sheet 1: Ringkasan"""
        ws = wb.create_sheet("Ringkasan", 0)
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        
        # Title
        ws['A1'] = "HASIL UJIAN"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Exam info
        ws['A3'] = "Nama Ujian:"
        ws['B3'] = exam.get('title', 'N/A')
        ws['A4'] = "Tanggal:"
        ws['B4'] = exam.get('date', 'N/A')
        ws['A5'] = "Mata Pelajaran:"
        ws['B5'] = exam.get('subject', 'N/A')
        ws['A6'] = "Kelas:"
        ws['B6'] = exam.get('class_name', exam.get('class', 'N/A'))
        ws['A7'] = "Jumlah Soal:"
        ws['B7'] = exam.get('active_questions', 0)
        
        # Statistics
        ws['A9'] = "STATISTIK"
        ws['A9'].font = Font(bold=True, size=12)
        
        ws['A10'] = "Total Siswa:"
        ws['B10'] = stats['total_students']
        ws['A11'] = "Nilai Rata-rata:"
        ws['B11'] = f"{stats['average_score']:.2f}"
        ws['A12'] = "Nilai Tertinggi:"
        ws['B12'] = f"{stats['highest_score']:.2f}"
        ws['A13'] = "Nilai Terendah:"
        ws['B13'] = f"{stats['lowest_score']:.2f}"
        ws['A14'] = "Tingkat Kelulusan:"
        ws['B14'] = f"{stats['pass_rate']:.2f}%"
        
        # Format columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_scores_sheet(self, wb, exam, results):
        """Sheet 2: Nilai Siswa"""
        ws = wb.create_sheet("Nilai Siswa")
        
        # Header
        headers = ['No', 'Nama Siswa', 'No. Induk', 'Benar', 'Salah', 'Kosong', 'Skor', 'Nilai Siswa', 'Predikat']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for idx, result in enumerate(results, start=2):
            # Calculate nilai (score out of 100)
            nilai_siswa = round(result['score']['percentage'], 2)
            
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=result.get('student_name', 'N/A'))
            ws.cell(row=idx, column=3, value=result.get('student_number', 'N/A'))
            ws.cell(row=idx, column=4, value=result['score']['correct'])
            ws.cell(row=idx, column=5, value=result['score']['wrong'])
            ws.cell(row=idx, column=6, value=result['score']['unanswered'])
            ws.cell(row=idx, column=7, value=result['score']['correct'])
            ws.cell(row=idx, column=8, value=nilai_siswa)
            ws.cell(row=idx, column=9, value=self._get_predicate(result['score']['percentage']))
        
        # Format columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_details_sheet(self, wb, exam, results):
        """Sheet 3: Detail Per Soal"""
        ws = wb.create_sheet("Detail Per Soal")
        
        # Header
        ws.cell(row=1, column=1, value="No Soal")
        ws.cell(row=1, column=2, value="Kunci")
        
        # Student columns
        for idx, result in enumerate(results, start=3):
            student_name = result.get('student_name', f'Siswa {idx-2}')
            ws.cell(row=1, column=idx, value=student_name[:15])
        
        # Data per question
        for q_num in range(exam['active_questions']):
            ws.cell(row=q_num+2, column=1, value=q_num+1)
            
            # Answer key
            key_idx = exam['answer_key'].get(str(q_num), -1)
            key_letter = chr(65 + key_idx) if key_idx >= 0 else '?'
            ws.cell(row=q_num+2, column=2, value=key_letter)
            
            # Student answers
            for idx, result in enumerate(results, start=3):
                ans_idx = result['answers'].get(str(q_num), -1)
                ans_letter = chr(65 + ans_idx) if ans_idx >= 0 else '-'
                
                cell = ws.cell(row=q_num+2, column=idx, value=ans_letter)
                
                # Color code
                if ans_idx == key_idx and ans_idx >= 0:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                elif ans_idx != -1:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
    
    def _get_predicate(self, percentage: float) -> str:
        """Convert percentage to grade predicate"""
        if percentage >= 90:
            return "A (Sangat Baik)"
        elif percentage >= 80:
            return "B (Baik)"
        elif percentage >= 70:
            return "C (Cukup)"
        elif percentage >= 60:
            return "D (Kurang)"
        else:
            return "E (Sangat Kurang)"
